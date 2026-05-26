from __future__ import annotations

from datetime import datetime, time
from pathlib import Path
from typing import Any, Dict, List, Optional

from core.base_context import event_anchor_by_key, load_base_context, section_anchor_by_key
from core.scenario_config import load_scenarios_for_config
from core.types import (
    AnalyzeConfig,
    AppConfig,
    BaseContext,
    BuildConfig,
    DelayScenario,
    EventAnchor,
    ExportTimetableConfig,
    InputConfig,
    ProjectConfig,
    RawTable,
    ScenarioConfig,
    SectionAnchor,
    SolveConfig,
    SolverConfig,
    SpeedLimitScenario,
)

TIME_HEADERS = {"arrival_time", "departure_time"}


def _require_yaml() -> Any:
    try:
        import yaml
    except ImportError as exc:  # pragma: no cover
        raise RuntimeError("Missing dependency: pyyaml") from exc
    return yaml


def _normalize_header(header: str) -> str:
    return header.lower()


def _normalize_time_text(value: str) -> str:
    parts = value.split(":")
    if len(parts) != 3 or any(part == "" for part in parts):
        return value
    try:
        hour, minute, second = [int(part) for part in parts]
    except ValueError:
        return value
    if hour < 0 or hour > 23 or minute < 0 or minute > 59 or second < 0 or second > 59:
        return value
    return f"{hour:02d}:{minute:02d}:{second:02d}"


def _normalize_cell_value(cell: Any, header: str) -> Optional[str]:
    if cell is None:
        return None
    if isinstance(cell, datetime):
        cell = cell.time()
    if isinstance(cell, time):
        return cell.strftime("%H:%M:%S")
    text = str(cell).strip()
    if text == "":
        return None
    if header in TIME_HEADERS:
        return _normalize_time_text(text)
    return text


def _read_excel(path: Path, sheet_name: str) -> RawTable:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover
        raise RuntimeError("Missing dependency: openpyxl") from exc

    workbook = load_workbook(path, read_only=True, data_only=True)
    if sheet_name not in workbook.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' not found in {path}")
    worksheet = workbook[sheet_name]
    rows = list(worksheet.iter_rows(values_only=True))
    if not rows:
        raise ValueError(f"Empty sheet: {path}#{sheet_name}")

    raw_headers = [_normalize_header(str(cell).strip()) if cell is not None else "" for cell in rows[0]]
    headers: List[str] = []
    header_indexes: List[int] = []
    for index, header in enumerate(raw_headers):
        if header == "":
            continue
        if header in headers:
            raise ValueError(f"Duplicated header in {path}#{sheet_name}: {header}")
        headers.append(header)
        header_indexes.append(index)

    records: List[Dict[str, Optional[str]]] = []
    for row in rows[1:]:
        if all(cell is None for cell in row):
            continue
        record: Dict[str, Optional[str]] = {}
        for header, index in zip(headers, header_indexes):
            cell = row[index] if index < len(row) else None
            record[header] = _normalize_cell_value(cell, header)
        records.append(record)
    return RawTable(headers=headers, rows=records)


def _parse_time_to_seconds(value: Any) -> int:
    # YAML may parse unquoted "HH:MM:SS" into a numeric scalar (e.g. 08:00:00 -> 28800).
    # Support both explicit "HH:MM:SS" strings and numeric seconds.
    if isinstance(value, (int, float)):
        seconds = int(value)
        if seconds < 0 or seconds > 24 * 3600 - 1:
            raise ValueError(f"Invalid HH:MM:SS time: {value}")
        return seconds

    text = str(value).strip()
    if text.isdigit():
        seconds = int(text)
        if seconds < 0 or seconds > 24 * 3600 - 1:
            raise ValueError(f"Invalid HH:MM:SS time: {value}")
        return seconds

    parts = text.split(":")
    if len(parts) != 3:
        raise ValueError(f"Invalid HH:MM:SS time: {value}")
    hour, minute, second = [int(part) for part in parts]
    if hour < 0 or hour > 23 or minute < 0 or minute > 59 or second < 0 or second > 59:
        raise ValueError(f"Invalid HH:MM:SS time: {value}")
    return hour * 3600 + minute * 60 + second


def _path_or_default(value: Any, default: Path) -> Path:
    if value is None:
        return default
    text = str(value).strip()
    if text == "":
        return default
    return Path(text)


def _str_or_default(value: Any, default: str) -> str:
    if value is None:
        return default
    text = str(value).strip()
    return text or default


def _bool_or_default(value: Any, default: bool) -> bool:
    if value is None:
        return default
    if isinstance(value, bool):
        return value
    text = str(value).strip().lower()
    if text in {"1", "true", "yes", "y", "on"}:
        return True
    if text in {"0", "false", "no", "n", "off"}:
        return False
    return default


def _required_path(value: Any, field_name: str) -> Path:
    if value is None:
        raise ValueError(f"Missing required config field: {field_name}")
    text = str(value).strip()
    if text == "":
        raise ValueError(f"Missing required config field: {field_name}")
    return Path(text)


def _has_text(item: Dict[str, Any], key: str) -> bool:
    return key in item and str(item.get(key, "")).strip() != ""


def _resolve_event_anchor(item: Dict[str, Any], base_context: BaseContext) -> EventAnchor:
    anchor_id = str(item.get("event_anchor_id", "")).strip()
    anchor = base_context.event_anchors.get(anchor_id) if anchor_id else None
    if anchor_id and anchor is None:
        raise ValueError(f"Unknown event_anchor_id: {anchor_id}")

    semantic_keys = ("train_id", "station", "event_type")
    semantic_present = [_has_text(item, key) for key in semantic_keys]
    if any(semantic_present):
        missing = [key for key, present in zip(semantic_keys, semantic_present) if not present]
        if missing:
            raise ValueError(
                "Delay scenarios using semantic fields must include train_id, station, and event_type; "
                f"missing: {', '.join(missing)}"
            )
        key = (
            str(item["train_id"]).strip(),
            str(item["station"]).strip(),
            str(item["event_type"]).strip(),
        )
        semantic_anchor = event_anchor_by_key(base_context).get(key)
        if semantic_anchor is None:
            raise ValueError(f"Delay scenario event not found in BaseContext: {key}")
        if anchor is not None and anchor.anchor_id != semantic_anchor.anchor_id:
            raise ValueError(
                "Delay scenario event_anchor_id conflicts with train_id/station/event_type: "
                f"{anchor.anchor_id} != {semantic_anchor.anchor_id}"
            )
        return semantic_anchor

    if anchor is None:
        raise ValueError(
            "Delay scenarios must use event_anchor_id or train_id/station/event_type."
        )
    return anchor


def _resolve_section_anchor(item: Dict[str, Any], base_context: BaseContext) -> SectionAnchor:
    anchor_id = str(item.get("section_anchor_id", "")).strip()
    anchor = base_context.section_anchors.get(anchor_id) if anchor_id else None
    if anchor_id and anchor is None:
        raise ValueError(f"Unknown section_anchor_id: {anchor_id}")

    semantic_keys = ("start_station", "end_station")
    semantic_present = [_has_text(item, key) for key in semantic_keys]
    if any(semantic_present):
        missing = [key for key, present in zip(semantic_keys, semantic_present) if not present]
        if missing:
            raise ValueError(
                "Speed limit scenarios using semantic fields must include start_station and end_station; "
                f"missing: {', '.join(missing)}"
            )
        key = (
            str(item["start_station"]).strip(),
            str(item["end_station"]).strip(),
        )
        semantic_anchor = section_anchor_by_key(base_context).get(key)
        if semantic_anchor is None:
            raise ValueError(f"Speed limit scenario section not found in BaseContext: {key}")
        if anchor is not None and anchor.anchor_id != semantic_anchor.anchor_id:
            raise ValueError(
                "Speed limit scenario section_anchor_id conflicts with start_station/end_station: "
                f"{anchor.anchor_id} != {semantic_anchor.anchor_id}"
            )
        return semantic_anchor

    if anchor is None:
        raise ValueError(
            "Speed limit scenarios must use section_anchor_id or start_station/end_station."
        )
    return anchor


def load_config(path: Path) -> AppConfig:
    yaml = _require_yaml()
    with path.open("r", encoding="utf-8") as file:
        payload = yaml.safe_load(file) or {}

    project_cfg = payload.get("project", {}) or {}
    build_cfg = payload.get("build", {}) or {}
    solve_cfg = payload.get("solve", {}) or {}
    export_cfg = payload.get("export-timetable", payload.get("export_timetable", {})) or {}
    analyze_cfg = payload.get("analyze", payload.get("analysis", {})) or {}

    if "input" in payload:
        raise ValueError("Legacy config section 'input' is no longer supported; use project.base_context_path.")
    root_solver_cfg = payload.get("solver", {}) or {}
    if "scenarios" in payload:
        raise ValueError("Legacy top-level 'scenarios' is no longer supported; use build.scenarios.")
    for legacy_key in ("timetable_path", "mileage_path", "timetable_sheet_name", "mileage_sheet_name"):
        if legacy_key in project_cfg:
            raise ValueError(f"Legacy project.{legacy_key} is no longer supported; use project.base_context_path.")

    case_name = _str_or_default(project_cfg.get("name"), path.stem)
    output_dir = _path_or_default(
        project_cfg.get("output_dir"),
        Path("outputs") / "main" / "datasets" / case_name / "cases" / case_name,
    )
    base_context_path = _required_path(project_cfg.get("base_context_path"), "project.base_context_path")
    base_context = load_base_context(base_context_path)

    # Convention-first artifact paths.
    # build output: output_dir/<name>.lp
    # solve output: output_dir/<name>.sol
    # export output: output_dir/adjusted_timetable.xlsx
    lp_default_path = output_dir / f"{case_name}.lp"
    solution_default_path = output_dir / f"{case_name}.sol"
    adjusted_default_path = output_dir / "adjusted_timetable.xlsx"
    metrics_default_path = output_dir / "analysis_metrics.xlsx"
    plot_default_path = output_dir / "timetable_plot.png"

    scenarios_cfg = load_scenarios_for_config(build_cfg.get("scenarios", {}) or {}, path, yaml)
    if "interruptions" in scenarios_cfg:
        raise ValueError("Legacy build.scenarios.interruptions is no longer supported; use speed_limits with limit_speed=0.")

    solve_solver_cfg = solve_cfg.get("solver", {}) or {}

    def _solve_value(key: str, default: Any) -> Any:
        if key in solve_cfg:
            return solve_cfg[key]
        if key in solve_solver_cfg:
            return solve_solver_cfg[key]
        return root_solver_cfg.get(key, default)

    solve_lp_path = _path_or_default(solve_cfg.get("lp_path"), lp_default_path)
    export_solution_path = _path_or_default(
        export_cfg.get("sol_path", export_cfg.get("solution_path")),
        solution_default_path,
    )

    adjusted_timetable_path = _path_or_default(
        analyze_cfg.get("adj_timetable_path", analyze_cfg.get("adjusted_timetable_path")),
        adjusted_default_path,
    )
    adjusted_timetable_sheet_name = _str_or_default(
        analyze_cfg.get("adj_timetable_sheet_name", analyze_cfg.get("adjusted_timetable_sheet_name")),
        "Sheet1",
    )
    metrics_output_path = _path_or_default(analyze_cfg.get("metrics_output_path"), metrics_default_path)
    plot_output_path = _path_or_default(analyze_cfg.get("plot_output_path"), plot_default_path)

    delays = []
    for item in scenarios_cfg.get("delays", []):
        anchor = _resolve_event_anchor(item, base_context)
        seconds = int(item["seconds"])
        if seconds <= 0:
            raise ValueError("Delay seconds must be > 0")
        delays.append(
            DelayScenario(
                event_anchor_id=anchor.anchor_id,
                train_id=anchor.train_id,
                station=anchor.station,
                event_type=anchor.event_type,
                seconds=seconds,
            )
        )

    speed_limits = []
    for item in scenarios_cfg.get("speed_limits", []):
        if any(key in item for key in ("extra_seconds", "end_time")):
            raise ValueError(
                "Speed limit scenarios must use section_anchor_id/start_time/duration/limit_speed; "
                "extra_seconds/end_time are no longer supported."
            )
        anchor = _resolve_section_anchor(item, base_context)
        start_time = _parse_time_to_seconds(item["start_time"])
        duration = int(item["duration"])
        limit_speed = float(item["limit_speed"])
        if duration <= 0:
            raise ValueError("Speed limit duration must be > 0")
        if start_time + duration > 24 * 3600:
            raise ValueError("Speed limit start_time + duration must not exceed 24:00:00")
        if limit_speed < 0:
            raise ValueError("Speed limit limit_speed must be >= 0")
        speed_limits.append(
            SpeedLimitScenario(
                section_anchor_id=anchor.anchor_id,
                start_station=anchor.start_station,
                end_station=anchor.end_station,
                start_time=start_time,
                duration=duration,
                limit_speed=limit_speed,
            )
        )

    objective_mode_raw = str(_solve_value("objective_mode", "abs")).strip()
    cancellation_default = False
    objective_mode = objective_mode_raw
    if objective_mode_raw == "cal_delay_plus_cancel":
        # Compatibility alias: old mixed mode now maps to
        # abs objective + independent cancellation switch.
        objective_mode = "abs"
        cancellation_default = True

    return AppConfig(
        project=ProjectConfig(name=case_name, output_dir=output_dir, base_context_path=base_context_path),
        input=InputConfig(
            timetable_path=base_context.source_timetable_path,
            mileage_path=base_context.source_mileage_path,
            timetable_sheet_name=base_context.timetable_sheet_name,
            mileage_sheet_name=base_context.mileage_sheet_name,
        ),
        solver=SolverConfig(
            objective_delay_weight=float(_solve_value("objective_delay_weight", 1.0)),
            objective_mode=objective_mode,
            cancellation_enabled=_bool_or_default(
                _solve_value("cancellation_enabled", cancellation_default),
                cancellation_default,
            ),
            cancellation_penalty_weight=float(
                _solve_value("cancellation_penalty_weight", 1000.0)
            ),
            arr_arr_headway_seconds=int(_solve_value("arr_arr_headway_seconds", 180)),
            dep_dep_headway_seconds=int(_solve_value("dep_dep_headway_seconds", 180)),
            dwell_seconds_at_stops=int(_solve_value("dwell_seconds_at_stops", 120)),
            big_m=int(_solve_value("big_m", 100000)),
            tolerance_delay_seconds=int(
                _solve_value(
                    "cancellation_threshold_seconds",
                    _solve_value("tolerance_delay_seconds", 2 * 3600),
                )
            ),
        ),
        scenarios=ScenarioConfig(
            delays=delays,
            speed_limits=speed_limits,
        ),
        build=BuildConfig(lp_path=lp_default_path),
        solve=SolveConfig(lp_path=solve_lp_path, solution_path=solution_default_path),
        export_timetable=ExportTimetableConfig(
            solution_path=export_solution_path,
            timetable_path=adjusted_default_path,
        ),
        analyze=AnalyzeConfig(
            enable_metrics=bool(analyze_cfg.get("enable_metrics", False)),
            enable_plot=bool(analyze_cfg.get("enable_plot", False)),
            plot_grid=bool(analyze_cfg.get("plot_grid", False)),
            plot_title=str(analyze_cfg.get("plot_title", "Train Timetable")),
            plan_timetable_path=base_context.source_timetable_path,
            plan_timetable_sheet_name=base_context.timetable_sheet_name,
            adjusted_timetable_path=adjusted_timetable_path,
            adjusted_timetable_sheet_name=adjusted_timetable_sheet_name,
            metrics_output_path=metrics_output_path,
            plot_output_path=plot_output_path,
            plot_timetable_path=adjusted_timetable_path,
        ),
        base_context=base_context,
    )


def load_timetable(path: Path, sheet_name: str) -> RawTable:
    return _read_excel(path, sheet_name)


def load_mileage_table(path: Path, sheet_name: str) -> RawTable:
    return _read_excel(path, sheet_name)
